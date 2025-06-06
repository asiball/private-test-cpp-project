"""
プロジェクトドキュメント（Excel/CSV）生成スクリプト
実行: python3 tools/gen_docs.py
"""

import csv
import os
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── スタイル定数 ────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F497D")
SUB_FILL   = PatternFill("solid", fgColor="4F81BD")
ALT_FILL   = PatternFill("solid", fgColor="DCE6F1")
PASS_FILL  = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL  = PatternFill("solid", fgColor="FFC7CE")
SKIP_FILL  = PatternFill("solid", fgColor="FFEB9C")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT  = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Meiryo UI", size=10)
BOLD_FONT = Font(name="Meiryo UI", bold=True, size=10)

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _h(ws, row, col, value, width=None):
    """ヘッダセルを書き込む"""
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = HDR_FILL
    c.font   = HDR_FONT
    c.border = BORDER
    c.alignment = CENTER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _b(ws, row, col, value, fill=None, align=LEFT):
    """ボディセルを書き込む"""
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = fill or (ALT_FILL if row % 2 == 0 else WHITE_FILL)
    c.font   = BODY_FONT
    c.border = BORDER
    c.alignment = align
    return c


def _result_fill(result):
    if result == "○":   return PASS_FILL
    if result == "×":   return FAIL_FILL
    if result == "－":   return SKIP_FILL
    return WHITE_FILL


# ════════════════════════════════════════════════════════════
# 1. WBS
# ════════════════════════════════════════════════════════════
def gen_wbs():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WBS"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    headers = [
        ("ID", 6), ("フェーズ／タスク", 32), ("担当者", 10),
        ("開始日", 12), ("終了日", 12), ("工数(h)", 9),
        ("進捗(%)", 9), ("備考", 28),
    ]
    for col, (h, w) in enumerate(headers, 1):
        _h(ws, 1, col, h, w)

    tasks = [
        # ID, タスク名, 担当, 開始, 終了, 工数, 進捗, 備考
        ("1",   "■ 要件定義フェーズ",           "",     "2025-01-06", "2025-01-17", "",   "",    ""),
        ("1.1", "  要件ヒアリング",              "山田", "2025-01-06", "2025-01-10", 16,   100,   "発注者との打合せ"),
        ("1.2", "  要件定義書作成・レビュー",    "山田", "2025-01-13", "2025-01-17", 24,   100,   ""),
        ("2",   "■ 基本設計フェーズ",           "",     "2025-01-20", "2025-02-14", "",   "",    ""),
        ("2.1", "  システム構成設計",            "山田", "2025-01-20", "2025-01-24", 24,   100,   ""),
        ("2.2", "  インターフェース定義",        "山田", "2025-01-27", "2025-01-31", 16,   100,   "SPI IF確定"),
        ("2.3", "  基本設計書作成・レビュー",    "山田", "2025-02-03", "2025-02-14", 32,   100,   ""),
        ("3",   "■ 詳細設計フェーズ",           "",     "2025-02-17", "2025-03-14", "",   "",    ""),
        ("3.1", "  driverクラス設計",            "山田", "2025-02-17", "2025-02-21", 24,   100,   ""),
        ("3.2", "  libクラス設計",               "山田", "2025-02-24", "2025-02-28", 24,   100,   ""),
        ("3.3", "  API仕様書作成",               "山田", "2025-03-03", "2025-03-07", 16,   100,   ""),
        ("3.4", "  詳細設計書レビュー",          "山田", "2025-03-10", "2025-03-14", 16,   100,   ""),
        ("4",   "■ 実装フェーズ",               "",     "2025-03-17", "2025-05-09", "",   "",    ""),
        ("4.1", "  SpiDriver実装",               "山田", "2025-03-17", "2025-03-28", 40,   100,   ""),
        ("4.2", "  Device lib実装",              "山田", "2025-03-31", "2025-04-11", 40,   100,   ""),
        ("4.3", "  device-ctl CLI実装",          "山田", "2025-04-14", "2025-04-18", 24,   100,   ""),
        ("4.4", "  単体テスト実施",              "山田", "2025-04-21", "2025-04-25", 24,   100,   ""),
        ("4.5", "  EAGAIN修正（#42）",           "山田", "2025-03-17", "2025-03-19", 8,    100,   "不具合対応"),
        ("4.6", "  非同期API追加（#55）",        "山田", "2025-04-28", "2025-05-09", 32,   100,   ""),
        ("5",   "■ テストフェーズ",             "",     "2025-05-12", "2025-05-30", "",   "",    ""),
        ("5.1", "  結合テスト実施",              "山田", "2025-05-12", "2025-05-23", 40,   100,   ""),
        ("5.2", "  受け入れテスト（発注者立会）","山田", "2025-05-26", "2025-05-30", 16,   100,   ""),
        ("6",   "■ 納品フェーズ",               "",     "2025-06-02", "2025-06-06", "",   "",    ""),
        ("6.1", "  ドキュメント整備",            "山田", "2025-06-02", "2025-06-04", 16,   100,   ""),
        ("6.2", "  納品物確認・提出",            "山田", "2025-06-05", "2025-06-06", 8,    100,   ""),
    ]

    for r, (tid, name, person, start, end, effort, progress, note) in enumerate(tasks, 2):
        is_phase = name.startswith("■")
        fill = PatternFill("solid", fgColor="1F497D") if is_phase else None
        font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10) if is_phase else BODY_FONT

        for col, val in enumerate([tid, name, person, start, end, effort, progress, note], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill   = fill or (ALT_FILL if r % 2 == 0 else WHITE_FILL)
            c.font   = font
            c.border = BORDER
            c.alignment = CENTER if col in (1, 4, 5, 6, 7) else LEFT
            if col == 7 and isinstance(val, int):
                c.number_format = "0%"
                c.value = val / 100

    path = os.path.join(BASE, "docs/00_project/wbs.xlsx")
    wb.save(path)
    print(f"  created: {path}")


# ════════════════════════════════════════════════════════════
# 2. 単体テスト仕様書（driver）
# ════════════════════════════════════════════════════════════
def gen_unit_spec_driver():
    wb = openpyxl.Workbook()

    # ── シート1: テスト仕様 ──
    ws = wb.active
    ws.title = "テスト仕様"
    ws.sheet_view.showGridLines = False

    headers = [
        ("テストID", 12), ("対象クラス/メソッド", 24), ("試験区分", 10),
        ("試験条件／入力", 36), ("期待結果", 36), ("合否基準", 12),
    ]
    for col, (h, w) in enumerate(headers, 1):
        _h(ws, 1, col, h, w)

    specs = [
        ("UT-DRV-001", "SpiDriver::open()",    "正常系", "/dev/spidev0.0 が存在し、設定値が有効",       "戻り値 true, is_open()==true",                "完全一致"),
        ("UT-DRV-002", "SpiDriver::open()",    "異常系", "存在しないデバイスパス /dev/spidevXX.0",       "戻り値 false, last_errno() != 0",             "完全一致"),
        ("UT-DRV-003", "SpiDriver::close()",   "正常系", "open() 成功後に close() を呼ぶ",              "is_open()==false",                           "完全一致"),
        ("UT-DRV-004", "SpiDriver::close()",   "正常系", "open() せず close() を呼ぶ（二重close）",     "例外なし、クラッシュなし",                   "異常なし"),
        ("UT-DRV-005", "SpiDriver::transfer()", "正常系", "tx/rx 各8バイト、loopback接続",              "戻り値 8、rx==tx",                           "完全一致"),
        ("UT-DRV-006", "SpiDriver::transfer()", "異常系", "fd_ が未openの状態で transfer() を呼ぶ",     "戻り値 -1",                                  "完全一致"),
        ("UT-DRV-007", "SpiDriver::transfer()", "異常系", "EAGAIN を3回返すモックfd",                   "3回リトライ後に-1を返す",                    "完全一致"),
        ("UT-DRV-008", "SpiDriver（コピー禁止）", "正常系", "コピーコンストラクタ・代入演算子を呼ぼうとする", "コンパイルエラーとなること",              "コンパイル確認"),
    ]

    for r, row in enumerate(specs, 2):
        for col, val in enumerate(row, 1):
            _b(ws, r, col, val)

    # ── シート2: テスト結果 ──
    ws2 = wb.create_sheet("テスト結果")
    ws2.sheet_view.showGridLines = False

    hdrs2 = [
        ("テストID", 12), ("試験実施日", 14), ("実施者", 10),
        ("結果", 8), ("障害番号", 12), ("備考", 40),
    ]
    for col, (h, w) in enumerate(hdrs2, 1):
        _h(ws2, 1, col, h, w)

    results = [
        ("UT-DRV-001", "2025-04-22", "山田", "○", "",     ""),
        ("UT-DRV-002", "2025-04-22", "山田", "○", "",     ""),
        ("UT-DRV-003", "2025-04-22", "山田", "○", "",     ""),
        ("UT-DRV-004", "2025-04-22", "山田", "○", "",     ""),
        ("UT-DRV-005", "2025-04-22", "山田", "○", "",     "loopback治具使用"),
        ("UT-DRV-006", "2025-04-22", "山田", "○", "",     ""),
        ("UT-DRV-007", "2025-04-23", "山田", "○", "BUG-042", "初回はNG→EAGAIN修正(#42)後に再試験で○"),
        ("UT-DRV-008", "2025-04-23", "山田", "○", "",     "static_assertで確認"),
    ]

    for r, row in enumerate(results, 2):
        for col, val in enumerate(row, 1):
            fill = _result_fill(val) if col == 4 else None
            _b(ws2, r, col, val, fill=fill)

    path = os.path.join(BASE, "docs/06_test/unit/spec/driver-unit-spec.xlsx")
    wb.save(path)
    print(f"  created: {path}")


# ════════════════════════════════════════════════════════════
# 3. 単体テスト仕様書（lib）
# ════════════════════════════════════════════════════════════
def gen_unit_spec_lib():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "テスト仕様"
    ws.sheet_view.showGridLines = False

    headers = [
        ("テストID", 12), ("対象クラス/メソッド", 24), ("試験区分", 10),
        ("試験条件／入力", 36), ("期待結果", 36), ("合否基準", 12),
    ]
    for col, (h, w) in enumerate(headers, 1):
        _h(ws, 1, col, h, w)

    specs = [
        ("UT-LIB-001", "Device::open()",        "正常系", "有効なSPIパスを渡す",                        "戻り値 true, is_open()==true",           "完全一致"),
        ("UT-LIB-002", "Device::open()",        "異常系", "無効なSPIパスを渡す",                        "戻り値 false",                          "完全一致"),
        ("UT-LIB-003", "Device::read()",        "正常系", "reg=0x00, len=4, loopback",                 "4バイトのvectorが返る",                  "サイズ一致"),
        ("UT-LIB-004", "Device::read()",        "異常系", "未open状態で read() を呼ぶ",                 "空vectorが返る",                         "size()==0"),
        ("UT-LIB-005", "Device::write()",       "正常系", "reg=0x10, data={0xAA,0xBB}",               "戻り値 true",                            "完全一致"),
        ("UT-LIB-006", "Device::write()",       "異常系", "未open状態で write() を呼ぶ",                "戻り値 false",                           "完全一致"),
        ("UT-LIB-007", "Device::read_async()",  "正常系", "コールバックが呼ばれ、データが返る",          "コールバック内 err==0, data.size()==4",  "完全一致"),
        ("UT-LIB-008", "Device::read_async()",  "異常系", "未open状態でread_async()",                   "コールバック内 err!=0, data 空",         "完全一致"),
        ("UT-LIB-009", "Device（コピー禁止）",  "正常系", "コピー/代入を試みる",                        "コンパイルエラー",                       "コンパイル確認"),
    ]

    for r, row in enumerate(specs, 2):
        for col, val in enumerate(row, 1):
            _b(ws, r, col, val)

    ws2 = wb.create_sheet("テスト結果")
    ws2.sheet_view.showGridLines = False

    hdrs2 = [
        ("テストID", 12), ("試験実施日", 14), ("実施者", 10),
        ("結果", 8), ("障害番号", 12), ("備考", 40),
    ]
    for col, (h, w) in enumerate(hdrs2, 1):
        _h(ws2, 1, col, h, w)

    results = [
        ("UT-LIB-001", "2025-04-24", "山田", "○", "",     "MockSpiDriverを使用"),
        ("UT-LIB-002", "2025-04-24", "山田", "○", "",     ""),
        ("UT-LIB-003", "2025-04-24", "山田", "○", "",     "loopback治具使用"),
        ("UT-LIB-004", "2025-04-24", "山田", "○", "",     ""),
        ("UT-LIB-005", "2025-04-24", "山田", "○", "",     ""),
        ("UT-LIB-006", "2025-04-24", "山田", "○", "",     ""),
        ("UT-LIB-007", "2025-04-25", "山田", "○", "",     "コールバック完了をcondition_variableで待機"),
        ("UT-LIB-008", "2025-04-25", "山田", "○", "",     ""),
        ("UT-LIB-009", "2025-04-25", "山田", "○", "",     ""),
    ]

    for r, row in enumerate(results, 2):
        for col, val in enumerate(row, 1):
            fill = _result_fill(val) if col == 4 else None
            _b(ws2, r, col, val, fill=fill)

    path = os.path.join(BASE, "docs/06_test/unit/spec/lib-unit-spec.xlsx")
    wb.save(path)
    print(f"  created: {path}")


# ════════════════════════════════════════════════════════════
# 4. 結合テスト仕様書
# ════════════════════════════════════════════════════════════
def gen_integration_spec():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "テスト仕様"
    ws.sheet_view.showGridLines = False

    headers = [
        ("テストID", 12), ("試験項目", 30), ("試験区分", 10),
        ("前提条件", 28), ("操作手順", 40), ("期待結果", 36), ("合否基準", 12),
    ]
    for col, (h, w) in enumerate(headers, 1):
        _h(ws, 1, col, h, w)

    specs = [
        ("IT-001", "SPI書き込み→読み出し一致確認",  "正常系",
         "Raspi4B + SPI loopback接続",
         "1. device-ctl write 0x00 0xDE 0xAD\n2. device-ctl read 0x00 2",
         "read結果が 'de ad' と一致",      "完全一致"),
        ("IT-002", "複数バイト連続転送",             "正常系",
         "Raspi4B + SPI loopback接続",
         "device-ctl read 0x00 8 を10回連続実行",
         "全10回で8バイトが正常に返る",     "エラー0件"),
        ("IT-003", "高速転送（1MHz）での安定性",      "正常系",
         "Raspi4B、SPI speed=1MHz",
         "device-ctl read 0x00 64 を100回実行",
         "全100回でタイムアウト・エラーなし", "エラー0件"),
        ("IT-004", "EAGAINリトライ動作確認",          "異常系",
         "高負荷状態（CPUストレステスト並行）",
         "stress --cpu 4 & device-ctl read 0x00 4 を50回実行",
         "全50回で正常応答（リトライにより回復）", "エラー0件"),
        ("IT-005", "非同期read + 同期write の同時実行", "正常系",
         "Raspi4B + loopback",
         "device-ctl --async read 0x00 4 と write 0x10 0x01 を並行実行",
         "両操作が完了し、データ破壊なし",  "完全一致"),
        ("IT-006", "デバイスオープン失敗時のエラー処理", "異常系",
         "SPI デバイスノードが存在しない状態",
         "device-ctl -d /dev/spidevXX.0 read 0x00 1",
         "エラーメッセージが表示され exit code != 0", "exit code確認"),
    ]

    for r, row in enumerate(specs, 2):
        for col, val in enumerate(row, 1):
            _b(ws, r, col, val)

    ws2 = wb.create_sheet("テスト結果")
    ws2.sheet_view.showGridLines = False

    hdrs2 = [
        ("テストID", 12), ("試験実施日", 14), ("実施者", 10), ("使用機材", 20),
        ("結果", 8), ("障害番号", 12), ("備考", 36),
    ]
    for col, (h, w) in enumerate(hdrs2, 1):
        _h(ws2, 1, col, h, w)

    results = [
        ("IT-001", "2025-05-13", "山田", "Raspi4B, loopback治具", "○", "",        ""),
        ("IT-002", "2025-05-13", "山田", "Raspi4B, loopback治具", "○", "",        ""),
        ("IT-003", "2025-05-14", "山田", "Raspi4B, loopback治具", "○", "",        "全100回エラー0件"),
        ("IT-004", "2025-05-14", "山田", "Raspi4B, stress-ng",   "○", "",        "最大2回リトライで回復"),
        ("IT-005", "2025-05-15", "山田", "Raspi4B, loopback治具", "○", "",        ""),
        ("IT-006", "2025-05-15", "山田", "Raspi4B",              "○", "",        "exit code=1 を確認"),
    ]

    for r, row in enumerate(results, 2):
        for col, val in enumerate(row, 1):
            fill = _result_fill(val) if col == 5 else None
            _b(ws2, r, col, val, fill=fill)

    path = os.path.join(BASE, "docs/06_test/integration/spec/integration-spec.xlsx")
    wb.save(path)
    print(f"  created: {path}")


# ════════════════════════════════════════════════════════════
# 5. 受け入れテスト仕様書（発注者署名欄付き）
# ════════════════════════════════════════════════════════════
def gen_acceptance_spec():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "受入テスト仕様・結果"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 28

    # タイトル行
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1, value="受け入れテスト仕様書・結果報告書")
    title.fill = PatternFill("solid", fgColor="1F497D")
    title.font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=14)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 32

    # メタ情報
    meta = [
        ("プロジェクト名", "embedded-device-suite v1.1.0"),
        ("発注者",         "〇〇株式会社"),
        ("受注者",         "△△エンジニアリング"),
        ("試験実施日",     "2025-05-26 〜 2025-05-30"),
        ("試験場所",       "発注者会議室（立会い実施）"),
    ]
    for i, (k, v) in enumerate(meta, 2):
        kc = ws.cell(row=i, column=1, value=k)
        kc.font = BOLD_FONT
        kc.fill = SUB_FILL
        kc.border = BORDER
        kc.alignment = CENTER
        ws.merge_cells(f"B{i}:F{i}")
        vc = ws.cell(row=i, column=2, value=v)
        vc.font = BODY_FONT
        vc.border = BORDER
        vc.alignment = LEFT

    # ヘッダ行
    headers = ["テストID", "確認項目", "確認手順", "合格基準", "結果", "備考"]
    for col, h in enumerate(headers, 1):
        _h(ws, 8, col, h)

    specs = [
        ("AT-001", "SPI基本読み書き動作",   "device-ctl write/readコマンドを実行し、データが一致することを目視確認する",        "送受信データが一致",       "○", "発注者立会い"),
        ("AT-002", "エラー時の動作",        "存在しないデバイスでコマンドを実行し、適切なエラーメッセージが表示されることを確認", "エラーメッセージが表示される", "○", ""),
        ("AT-003", "連続動作の安定性",      "100回連続でread/writeを実行し、全回正常完了することを確認",                     "エラー0件",                "○", "3分程度"),
        ("AT-004", "非同期API動作確認",     "--asyncオプションでread実行し、コールバックが呼ばれることを確認",                 "正常にデータが出力される",   "○", ""),
        ("AT-005", "ライブラリの動的リンク確認", "lddコマンドでlibdevice.soが正しくリンクされていることを確認",               "libdevice.so.1 が表示される", "○", ""),
        ("AT-006", "ドキュメント一式の納品確認", "納品物チェックリストに従い全ドキュメントの存在を確認する",                 "全項目 ✓",                  "○", ""),
    ]

    for r, row in enumerate(specs, 9):
        for col, val in enumerate(row, 1):
            fill = _result_fill(val) if col == 5 else None
            _b(ws, r, col, val, fill=fill)

    # 署名欄
    sig_row = 9 + len(specs) + 2
    ws.merge_cells(f"A{sig_row}:F{sig_row}")
    sh = ws.cell(row=sig_row, column=1, value="承認署名欄")
    sh.fill = PatternFill("solid", fgColor="4F81BD")
    sh.font = HDR_FONT
    sh.border = BORDER
    sh.alignment = CENTER

    sig_data = [
        ("発注者確認者氏名", ""),
        ("発注者確認者署名", ""),
        ("確認日",           "2025-05-30"),
        ("受注者担当者",     "山田 太郎"),
    ]
    for i, (k, v) in enumerate(sig_data, sig_row + 1):
        kc = ws.cell(row=i, column=1, value=k)
        kc.fill = ALT_FILL
        kc.font = BOLD_FONT
        kc.border = BORDER
        kc.alignment = CENTER
        ws.merge_cells(f"B{i}:F{i}")
        vc = ws.cell(row=i, column=2, value=v)
        vc.font = BODY_FONT
        vc.border = BORDER
        vc.alignment = LEFT
        ws.row_dimensions[i].height = 28

    path = os.path.join(BASE, "docs/06_test/acceptance/spec/acceptance-spec.xlsx")
    wb.save(path)
    print(f"  created: {path}")


# ════════════════════════════════════════════════════════════
# 6. 納品物チェックリスト
# ════════════════════════════════════════════════════════════
def gen_delivery_checklist():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "納品物チェックリスト"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value="納品物チェックリスト — embedded-device-suite v1.1.0")
    title.fill = PatternFill("solid", fgColor="1F497D")
    title.font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=13)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers = [("No.", 6), ("納品物", 44), ("形式", 14), ("格納場所（リポジトリパス）", 44), ("確認", 8)]
    for col, (h, w) in enumerate(headers, 1):
        _h(ws, 2, col, h, w)

    items = [
        # ドキュメント
        ("",  "── ドキュメント ──",                         "",       "",                                          ""),
        (1,   "要件定義書",                                 "Markdown", "docs/01_requirements/requirements-spec.md", "✓"),
        (2,   "仕様変更履歴",                               "CSV",    "docs/01_requirements/change-log.csv",       "✓"),
        (3,   "システム構成図",                             "Draw.io", "docs/02_basic-design/system-architecture.drawio", "✓"),
        (4,   "基本設計書",                                 "Markdown", "docs/02_basic-design/system-architecture.md", "✓"),
        (5,   "詳細設計書（driver）",                       "Markdown", "docs/03_detailed-design/driver-design.md","✓"),
        (6,   "詳細設計書（lib）",                          "Markdown", "docs/03_detailed-design/lib-design.md",   "✓"),
        (7,   "API仕様書（SpiDriver）",                     "Markdown", "docs/04_api-spec/spi-driver-api.md",      "✓"),
        (8,   "API仕様書（Device lib）",                    "Markdown", "docs/04_api-spec/libdevice-api.md",       "✓"),
        (9,   "SPIハードウェアIF仕様書",                    "Markdown", "docs/05_interface-spec/spi-hardware-if.md","✓"),
        (10,  "テスト計画書",                               "Markdown", "docs/06_test/test-plan.md",               "✓"),
        (11,  "単体テスト仕様書・結果（driver）",           "Excel",  "docs/06_test/unit/spec/driver-unit-spec.xlsx","✓"),
        (12,  "単体テスト仕様書・結果（lib）",              "Excel",  "docs/06_test/unit/spec/lib-unit-spec.xlsx", "✓"),
        (13,  "結合テスト仕様書・結果",                     "Excel",  "docs/06_test/integration/spec/integration-spec.xlsx","✓"),
        (14,  "受け入れテスト仕様書・結果（署名付き）",     "Excel",  "docs/06_test/acceptance/spec/acceptance-spec.xlsx","✓"),
        (15,  "リリースノート v1.1.0",                      "Markdown", "docs/07_delivery/release-notes/v1.1.0.md","✓"),
        # ソースコード
        ("",  "── ソースコード ──",                         "",       "",                                          ""),
        (16,  "SPIドライバソース",                          "C++",    "driver/",                                   "✓"),
        (17,  "共有ライブラリソース",                        "C++",    "lib/",                                      "✓"),
        (18,  "CLIツールソース",                            "C++",    "cli/",                                      "✓"),
        (19,  "単体テストコード",                           "C++",    "tests/unit/",                               "✓"),
        (20,  "結合テストコード",                           "C++",    "tests/integration/",                        "✓"),
        (21,  "ビルドスクリプト",                           "CMake",  "CMakeLists.txt, */CMakeLists.txt",          "✓"),
        (22,  "CI/CD設定",                                  "YAML",   ".gitlab-ci.yml",                           "✓"),
    ]

    for r, (no, name, fmt, path, check) in enumerate(items, 3):
        is_section = no == ""
        fill = PatternFill("solid", fgColor="4F81BD") if is_section else None
        font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10) if is_section else BODY_FONT
        for col, val in enumerate([no, name, fmt, path, check], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill   = fill or (ALT_FILL if r % 2 == 0 else WHITE_FILL)
            c.font   = font
            c.border = BORDER
            c.alignment = CENTER if col in (1, 5) else LEFT

    path = os.path.join(BASE, "docs/07_delivery/delivery-checklist.xlsx")
    wb.save(path)
    print(f"  created: {path}")


# ════════════════════════════════════════════════════════════
# 7. 仕様変更履歴（CSV）
# ════════════════════════════════════════════════════════════
def gen_change_log():
    rows = [
        ["変更ID", "日付",       "変更種別", "変更箇所",              "変更内容",                         "依頼者",   "対応者", "対応バージョン"],
        ["CHG-001","2025-01-17", "追加",    "要件定義書 §3.2",       "SPI speed の最大値を2MHzに変更",   "〇〇株式会社","山田", "driver/v1.0.0"],
        ["CHG-002","2025-03-10", "修正",    "詳細設計書 §4.1",       "EAGAIN時のリトライ仕様を追加",     "内部検討",  "山田",  "driver/v1.0.1"],
        ["CHG-003","2025-04-20", "追加",    "API仕様書 libdevice §2","非同期API read_async() を追加",    "〇〇株式会社","山田", "lib/v1.1.0"],
        ["CHG-004","2025-04-20", "追加",    "CLI仕様 §3",            "--async フラグを追加",             "〇〇株式会社","山田", "cli/v1.1.0"],
    ]
    path = os.path.join(BASE, "docs/01_requirements/change-log.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)
    print(f"  created: {path}")


# ════════════════════════════════════════════════════════════
# メイン
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Generating project documents...")
    gen_wbs()
    gen_unit_spec_driver()
    gen_unit_spec_lib()
    gen_integration_spec()
    gen_acceptance_spec()
    gen_delivery_checklist()
    gen_change_log()
    print("Done.")
